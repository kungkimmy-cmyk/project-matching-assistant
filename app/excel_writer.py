"""
excel_writer.py
----------------
Builds Matching_Results.xlsx. Updated for the new business rules:
theme/factory reasoning shown per project, best-match-plus-alternatives
per product (auto-selected vs needs-review), reactive-glaze flagging,
and an explicit internal-only marking on every sheet.

SECURITY NOTE (per spec: "customer-facing quotation output must never
expose factory cost, internal margin, bank information, or other
internal cost information"): every sheet this module writes is
explicitly an INTERNAL working document -- this tool does not generate
customer-facing quotations at all. A genuinely separate, redacted
customer-facing exporter (using cfg.internal_only_columns to drop
factory-cost/margin/bank columns) would be a distinct future
deliverable, not something silently bolted on here. Every sheet is
labeled accordingly so this is never ambiguous.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .orchestrator import MatchingRunResult

logger = logging.getLogger(__name__)

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2E4B6E", end_color="2E4B6E", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
CONFIRMED_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
UNMATCHED_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
REACTIVE_FILL = PatternFill(start_color="E6D6F5", end_color="E6D6F5", fill_type="solid")
SECURITY_BANNER_FONT = Font(name="Arial", bold=True, color="B00000", size=10)

SECURITY_BANNER = (
    "CONFIDENTIAL -- INTERNAL USE ONLY. Contains factory cost and internal pricing data. "
    "Do NOT send this file, or any data copied from it, to a customer."
)


def _write_security_banner(ws: Worksheet, row: int = 1) -> int:
    cell = ws.cell(row=row, column=1, value=SECURITY_BANNER)
    cell.font = SECURITY_BANNER_FONT
    return row + 2


def _write_table(ws: Worksheet, headers: List[str], rows: List[tuple], col_widths: List[int], row_fills=None, start_row: int = 1) -> None:
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = col_widths[idx - 1] if idx - 1 < len(col_widths) else 20
    for r_offset, row_values in enumerate(rows):
        r_idx = start_row + 1 + r_offset
        for c_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = BODY_FONT
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"
        if row_fills:
            fill = row_fills[r_offset] if r_offset < len(row_fills) else None
            if fill:
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill
    ws.freeze_panes = f"A{start_row + 1}"
    ws.row_dimensions[start_row].height = 28
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"


def _project_label(customer_file: str) -> str:
    return customer_file.rsplit(".", 1)[0]


def write_matching_results(result: MatchingRunResult, output_path: Path | str) -> Path:
    output_path = Path(output_path)
    wb = Workbook()
    m = result.matching
    analyses_by_pair = {(pa.match.factory_file, pa.match.customer_file): pa for pa in result.project_analyses}

    ws1 = wb.active
    ws1.title = "Confirmed Matches"
    start = _write_security_banner(ws1)
    rows1 = []
    for match in m.confirmed:
        pa = analyses_by_pair.get((match.factory_file, match.customer_file))
        theme_text = f"{pa.theme.dominant_theme} ({pa.theme.theme_confidence}%)" if pa else ""
        factory_rec = pa.factory_recommendation.factory if pa else ""
        products = pa.comparison_rows if pa else []
        rows1.append((
            match.factory_file, match.customer_file, match.confidence,
            "; ".join(match.reasons), match.factory_created or "", match.customer_created or "",
            factory_rec, theme_text, len(products), "",
        ))
    _write_table(
        ws1, ["Factory File", "Customer File", "Confidence", "Reasoning", "Factory Quotation Date",
              "Customer Quotation Date", "Recommended Factory", "RFQ Theme", "Products", "Notes"],
        rows1, [40, 40, 12, 60, 22, 22, 18, 26, 10, 30], start_row=start,
    )

    ws2 = wb.create_sheet("Needs Review")
    start2 = _write_security_banner(ws2)
    rows2 = [
        (match.factory_file, match.customer_file, match.confidence, "; ".join(match.reasons),
         match.factory_created or "", match.customer_created or "")
        for match in m.needs_review
    ]
    _write_table(
        ws2, ["Factory File", "Customer File", "Confidence", "Why It Needs Review",
              "Factory Quotation Date", "Customer Quotation Date"],
        rows2, [40, 40, 12, 80, 22, 22], start_row=start2,
    )

    ws3 = wb.create_sheet("Unmatched Factory Files")
    start3 = _write_security_banner(ws3)
    rows3 = []
    for fname in m.unmatched_factory:
        candidate = m.closest_candidate.get(fname)
        rows3.append((fname, candidate.customer_file if candidate else "", candidate.confidence if candidate else "",
                      "; ".join(candidate.reasons) if candidate else "No plausible candidate found"))
    _write_table(ws3, ["Factory File", "Closest Candidate (below threshold)", "Confidence", "Notes"], rows3, [45, 45, 14, 70], start_row=start3)

    ws4 = wb.create_sheet("Unmatched Customer Files")
    start4 = _write_security_banner(ws4)
    rows4 = []
    for fname in m.unmatched_customer:
        candidate = m.closest_candidate.get(fname)
        rows4.append((fname, candidate.factory_file if candidate else "", candidate.confidence if candidate else "",
                      "; ".join(candidate.reasons) if candidate else "No plausible candidate found"))
    _write_table(ws4, ["Customer File", "Closest Candidate (below threshold)", "Confidence", "Notes"], rows4, [45, 45, 14, 70], start_row=start4)

    ws5 = wb.create_sheet("Product Comparison")
    start5 = _write_security_banner(ws5)
    rows5, fills5 = [], []
    for pa in result.project_analyses:
        for p in pa.comparison_rows:
            alt_text = "; ".join(f"{a.db_code} ({a.confidence}%)" for a in p.alternatives) if p.alternatives else ""
            status = "Auto-selected" if p.auto_selected else ("Needs Review" if p.matched else "Unmatched")
            rows5.append((
                p.factory_code, p.db_code, p.description, p.factory_cost, p.selling_price,
                "Yes" if p.matched else "No", p.confidence, status, alt_text,
                "Yes" if p.reactive_glaze else "", p.historical_tier,
                "Yes" if p.historical_needs_review else "", p.theme_alignment_note, p.notes,
            ))
            fill = None
            if p.reactive_glaze:
                fill = REACTIVE_FILL
            elif not p.matched:
                fill = UNMATCHED_FILL
            elif not p.auto_selected:
                fill = REVIEW_FILL
            fills5.append(fill)
    _write_table(
        ws5,
        ["Factory Code", "DB Code", "Description", "Factory Cost", "Selling Price", "Matched", "Confidence",
         "Status", "Alternatives", "Reactive Glaze", "Historical Tier", "Historical Conflict",
         "Theme Alignment", "Notes"],
        rows5, [16, 18, 38, 13, 13, 9, 11, 15, 30, 12, 14, 14, 30, 40], fills5, start_row=start5,
    )

    ws6 = wb.create_sheet("DB Code Mapping")
    start6 = _write_security_banner(ws6)
    rows6, fills6 = [], []
    for mp in result.db_code_mapping:
        rows6.append((mp.factory_code, mp.primary_db_code, ", ".join(mp.alternative_db_codes),
                      mp.occurrences, ", ".join(mp.projects), mp.confidence, mp.recommendation))
        fills6.append(REVIEW_FILL if mp.alternative_db_codes else None)
    _write_table(
        ws6, ["Factory Code", "Primary DB Code", "Alternative DB Codes", "Occurrences", "Projects", "Confidence", "Recommendation"],
        rows6, [16, 18, 30, 12, 50, 24, 60], fills6, start_row=start6,
    )

    ws7 = wb.create_sheet("Master Migration File")
    start7 = _write_security_banner(ws7)
    rows7 = []
    for pa in result.project_analyses:
        project = _project_label(pa.match.customer_file)
        for p in pa.comparison_rows:
            status = "Matched (auto-selected)" if p.auto_selected else ("Matched (needs review)" if p.matched else "Unmatched (needs manual check)")
            source = pa.match.customer_file if p.db_code else pa.match.factory_file
            rows7.append((
                project, p.description, p.db_code, p.factory_code, p.selling_price, p.factory_cost,
                status, p.confidence, p.notes, source, "2025 Migration",
                "Yes" if p.reactive_glaze else "", p.historical_tier,
            ))
    _write_table(
        ws7,
        ["Project", "Description", "DB Brand Code", "Factory Code", "Quoted Price (USD)", "Factory Cost (RMB)",
         "Status", "Confidence Score", "Confidence Reasons", "Source File", "Source Folder",
         "Reactive Glaze", "Historical Tier"],
        rows7, [24, 40, 18, 16, 16, 16, 24, 14, 36, 30, 16, 12, 14], start_row=start7,
    )

    wb.save(output_path)
    logger.info("Wrote matching results to %s", output_path)
    return output_path
