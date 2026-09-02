"""
file_reader.py
--------------
Self-contained Excel reader for the Matching Assistant. Deliberately
NOT imported from the RFQ Master Database Extractor project -- this
is a completely separate application per the user's explicit request.

Produces a WorkbookSummary: document metadata, every sheet's cells
with fill colors, and (new) embedded image references per sheet for
photo_extractor.py to use.
"""
from __future__ import annotations

import logging
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)

LEGACY_EXTENSIONS = {".xls"}


class ReadError(RuntimeError):
    pass


@dataclass
class CellInfo:
    row: int
    col: int
    coordinate: str
    value: Any
    fill_rgb: Optional[str]


@dataclass
class ImageRef:
    """An embedded image, anchored near a given row/col in a sheet."""
    anchor_row: int
    anchor_col: int
    image_bytes: bytes
    format: str  # e.g. 'png', 'jpeg'


@dataclass
class SheetSummary:
    name: str
    max_row: int
    max_col: int
    cells: Dict[str, CellInfo] = field(default_factory=dict)
    images: List[ImageRef] = field(default_factory=list)

    def get(self, row: int, col: int) -> Optional[CellInfo]:
        return self.cells.get(f"{row},{col}")


@dataclass
class WorkbookSummary:
    path: Path
    filename: str
    created: Optional[str]
    modified: Optional[str]
    creator: Optional[str]
    sheets: List[SheetSummary] = field(default_factory=list)

    def all_text(self) -> str:
        parts = []
        for sheet in self.sheets:
            for cell in sheet.cells.values():
                if isinstance(cell.value, str):
                    parts.append(cell.value)
        return "\n".join(parts)


def _find_soffice_binary() -> Optional[str]:
    for name in ("soffice", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    return None


def _convert_legacy(path: Path, out_dir: Path) -> Path:
    binary = _find_soffice_binary()
    if not binary:
        raise ReadError(f"Cannot read legacy .xls ({path.name}): LibreOffice not installed.")
    out_dir.mkdir(parents=True, exist_ok=True)
    cmd = [binary, "--headless", "--norestore", "--convert-to", "xlsx", "--outdir", str(out_dir), str(path)]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    except (subprocess.TimeoutExpired, FileNotFoundError) as exc:
        raise ReadError(f"LibreOffice conversion failed for {path.name}: {exc}") from exc
    converted = out_dir / (path.stem + ".xlsx")
    if result.returncode != 0 or not converted.exists():
        raise ReadError(f"LibreOffice conversion failed for {path.name}: {result.stderr[:300]}")
    return converted


def _extract_fill_rgb(cell) -> Optional[str]:
    try:
        fill = cell.fill
        if fill is None or fill.fgColor is None:
            return None
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str) and rgb not in ("00000000",):
            return rgb
        return None
    except Exception:  # noqa: BLE001
        return None


def _extract_images(ws) -> List[ImageRef]:
    """openpyxl exposes embedded images via ws._images (each with an
    .anchor giving its row/col position). Best-effort: some workbooks
    have drawings openpyxl can't fully parse (see ConversionError
    handling upstream) -- any single image that fails to extract is
    skipped rather than aborting the whole sheet read."""
    images: List[ImageRef] = []
    raw_images = getattr(ws, "_images", None) or []
    for img in raw_images:
        try:
            anchor = img.anchor
            row = getattr(getattr(anchor, "_from", None), "row", 0) + 1
            col = getattr(getattr(anchor, "_from", None), "col", 0) + 1
            data = img._data() if callable(getattr(img, "_data", None)) else img.ref
            fmt = getattr(img, "format", None) or "png"
            images.append(ImageRef(anchor_row=row, anchor_col=col, image_bytes=data, format=str(fmt).lower()))
        except Exception as exc:  # noqa: BLE001
            logger.debug("Skipped one unreadable embedded image: %s", exc)
    return images


def read_workbook(
    path: Path | str, max_rows_per_sheet: int = 2000, max_cols_per_sheet: int = 60,
    extract_images: bool = True,
) -> WorkbookSummary:
    path = Path(path)
    tmp_ctx = None
    read_path = path
    if path.suffix.lower() in LEGACY_EXTENSIONS:
        tmp_ctx = tempfile.TemporaryDirectory(prefix="matcher_xls_")
        read_path = _convert_legacy(path, Path(tmp_ctx.name))

    try:
        wb_v = openpyxl.load_workbook(read_path, data_only=True, read_only=False)
        wb_f = openpyxl.load_workbook(read_path, data_only=False, read_only=False)

        created = modified = creator = None
        try:
            props = wb_v.properties
            created = props.created.isoformat() if props.created else None
            modified = props.modified.isoformat() if props.modified else None
            creator = props.creator
        except Exception:  # noqa: BLE001
            pass

        sheets: List[SheetSummary] = []
        for sheet_name in wb_v.sheetnames:
            ws_v = wb_v[sheet_name]
            ws_f = wb_f[sheet_name] if sheet_name in wb_f.sheetnames else ws_v
            max_row = min(ws_v.max_row or 0, max_rows_per_sheet)
            max_col = min(ws_v.max_column or 0, max_cols_per_sheet)
            cells: Dict[str, CellInfo] = {}
            for row in ws_v.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    value = cell.value
                    if value is None or (isinstance(value, str) and not value.strip()):
                        continue
                    r, c = cell.row, cell.column
                    fmt_cell = ws_f.cell(row=r, column=c)
                    cells[f"{r},{c}"] = CellInfo(
                        row=r, col=c, coordinate=cell.coordinate,
                        value=value, fill_rgb=_extract_fill_rgb(fmt_cell),
                    )
            images = _extract_images(ws_f) if extract_images else []
            sheets.append(SheetSummary(name=sheet_name, max_row=max_row, max_col=max_col, cells=cells, images=images))

        return WorkbookSummary(
            path=path, filename=path.name, created=created, modified=modified,
            creator=creator, sheets=sheets,
        )
    except ReadError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise ReadError(f"Failed to read {path.name}: {exc}") from exc
    finally:
        if tmp_ctx is not None:
            tmp_ctx.cleanup()
