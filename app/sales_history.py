"""
sales_history.py
------------------
Reads the Mimosa Sales Invoice Report -- the SOLD evidence source,
the strongest tier in the historical-evidence hierarchy (SOLD >
QUOTED > FACTORY QUOTED > CATALOGUE).

Column layout VALIDATED against the real uploaded file (not guessed):
Sales Invoice Date / No. / Customer Code / Product | Material Code /
Product | Material Brief Description / Quantity / Unit Code / Currency
Symbol / Unit Price / Disc. (%) / Amount. Read from matcher_config.
json's sales_report_column_map, so a differently-formatted future
export still just needs a config update, not a code change.

IMPORTANT, confirmed against the real file: there is NO separate
factory-code column here -- only "Product | Material Code" (the DB
code). See orchestrator.py for how a factory code is backfilled onto
SOLD records by cross-referencing the DB code against QUOTED/FACTORY
QUOTED records already read from the project files.

Missing file / missing columns degrade gracefully (empty SOLD
evidence, not a crash) rather than blocking the rest of the pipeline.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl

from .config import MatcherConfig

logger = logging.getLogger(__name__)


@dataclass
class SalesRecord:
    db_code: Optional[str]
    factory_code: Optional[str]  # not present in the real report; always None unless a future layout adds it
    description: Optional[str]
    quantity: Optional[float]
    unit_price: Optional[float]
    date: Optional[str]  # ISO date string where parseable, else the raw text
    customer: Optional[str]
    source_file: str
    invoice_no: Optional[str] = None
    amount: Optional[float] = None
    currency: Optional[str] = None


def _find_header_row(ws, column_names: List[str], max_scan_rows: int = 20) -> Optional[int]:
    """Sales/invoice exports commonly have a title block before the
    real header row -- scan the first N rows for one that contains at
    least half of the expected column names."""
    wanted = {c.lower() for c in column_names}
    for row in ws.iter_rows(min_row=1, max_row=max_scan_rows):
        found = {str(c.value).strip().lower() for c in row if isinstance(c.value, str)}
        if len(found & wanted) >= max(1, len(wanted) // 2):
            return row[0].row
    return None


def read_sales_history(path: Path | str, cfg: MatcherConfig) -> List[SalesRecord]:
    """Never raises for a missing/unreadable file -- returns an empty
    list and logs a warning, since SOLD evidence being unavailable
    should not block QUOTED-evidence-only matching."""
    path = Path(path)
    if not path.exists():
        logger.warning("Sales report not found at %s -- proceeding with QUOTED evidence only.", path)
        return []

    col_map = cfg.sales_report_column_map
    try:
        # NOTE: read_only=True previously caused a real bug on an actual
        # sales export -- openpyxl reported a stale 1x1 dimension for a
        # file that genuinely had 1,629 rows (the real <dimension> tag
        # was correct; read_only mode's fast-path didn't pick it up).
        # read_only=False costs more memory but reads correctly.
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not read sales report %s: %s -- proceeding with QUOTED evidence only.", path, exc)
        return []

    records: List[SalesRecord] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_num = _find_header_row(ws, list(col_map.values()))
        if header_row_num is None:
            continue

        header_row = next(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))
        col_index = {}
        for cell in header_row:
            if isinstance(cell.value, str):
                for field_name, header_label in col_map.items():
                    if cell.value.strip().lower() == header_label.strip().lower():
                        col_index[field_name] = cell.column

        if "db_code" not in col_index and "factory_code" not in col_index:
            continue  # this sheet doesn't look like the sales data (e.g. a notes/summary tab)

        for row in ws.iter_rows(min_row=header_row_num + 1):
            def _get(field_name):
                idx = col_index.get(field_name)
                if idx is None:
                    return None
                cell = row[idx - 1] if idx - 1 < len(row) else None
                return cell.value if cell else None

            db_code = _get("db_code")
            factory_code = _get("factory_code")
            if not db_code and not factory_code:
                continue

            date_val = _get("date")
            date_str = date_val.isoformat() if hasattr(date_val, "isoformat") else (str(date_val) if date_val else None)

            records.append(SalesRecord(
                db_code=str(db_code).strip().upper() if db_code else None,
                factory_code=str(factory_code).strip().upper() if factory_code else None,
                description=str(_get("description")).strip() if _get("description") else None,
                quantity=float(_get("quantity")) if isinstance(_get("quantity"), (int, float)) else None,
                unit_price=float(_get("unit_price")) if isinstance(_get("unit_price"), (int, float)) else None,
                date=date_str,
                customer=str(_get("customer")).strip() if _get("customer") else None,
                source_file=path.name,
                invoice_no=str(_get("invoice_no")).strip() if _get("invoice_no") else None,
                amount=float(_get("amount")) if isinstance(_get("amount"), (int, float)) else None,
                currency=str(_get("currency")).strip() if _get("currency") else None,
            ))

    logger.info("Read %d sales record(s) from %s", len(records), path.name)
    return records
