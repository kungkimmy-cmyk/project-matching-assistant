import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from openpyxl.styles import PatternFill

from app.orchestrator import run_matching
from app.config import MatcherConfig


def _make_workbook(path, db_code, factory_code="H0001"):
    wb = openpyxl.Workbook()
    ws = wb.active
    yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    ws.cell(row=1, column=1, value="Item Code")
    ws.cell(row=1, column=3, value="Factory Code")
    ws.cell(row=1, column=4, value="Unit Price (USD EXW)")
    ws.cell(row=1, column=3).fill = yellow
    ws.cell(row=2, column=1, value=db_code)
    ws.cell(row=2, column=2, value="plate")
    ws.cell(row=2, column=3, value=factory_code)
    ws.cell(row=2, column=3).fill = yellow
    ws.cell(row=2, column=4, value=1.0)
    wb.save(path)


class TestMultipleFoldersPerSide(unittest.TestCase):
    def setUp(self):
        self.root = Path(tempfile.mkdtemp())
        self.factory_a = self.root / "2025 Factory Quotations"
        self.factory_b = self.root / "2026 Factory Quotations"
        self.customer_a = self.root / "2025 Customer Quotations"
        self.customer_b = self.root / "2026 Customer Quotations"
        for d in (self.factory_a, self.factory_b, self.customer_a, self.customer_b):
            d.mkdir()
        _make_workbook(self.factory_a / "fa.xlsx", "DB000001", "H0001")
        _make_workbook(self.factory_b / "fb.xlsx", "DB000002", "H0002")
        _make_workbook(self.customer_a / "ca.xlsx", "DB000001", "H0001")
        _make_workbook(self.customer_b / "cb.xlsx", "DB000002", "H0002")
        self.cfg = MatcherConfig()

    def test_files_from_all_folders_are_combined(self):
        result = run_matching(
            [self.factory_a, self.factory_b], [self.customer_a, self.customer_b], self.cfg,
            photo_store_dir=self.root / "photo_store",
        )
        self.assertEqual(result.stats.factory_files_found, 2)
        self.assertEqual(result.stats.customer_files_found, 2)
        self.assertEqual(len(result.stats.factory_folders_scanned), 2)
        self.assertEqual(len(result.stats.customer_folders_scanned), 2)

    def test_single_folder_still_works_backward_compatible(self):
        # Old call style (bare path, not a list) must still work.
        result = run_matching(self.factory_a, self.customer_a, self.cfg, photo_store_dir=self.root / "photo_store")
        self.assertEqual(result.stats.factory_files_found, 1)
        self.assertEqual(result.stats.customer_files_found, 1)

    def test_duplicate_folder_in_list_does_not_double_count_files(self):
        result = run_matching(
            [self.factory_a, self.factory_a], [self.customer_a], self.cfg,
            photo_store_dir=self.root / "photo_store",
        )
        self.assertEqual(result.stats.factory_files_found, 1)  # de-duplicated


if __name__ == "__main__":
    unittest.main()
