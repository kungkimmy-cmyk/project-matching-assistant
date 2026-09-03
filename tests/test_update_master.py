"""
Integration tests for the complete Update Master flow. Covers, at
minimum, scenarios A-G as specified:
  A. first run on a folder
  B. immediate second run with no file changes
  C. one new file added
  D. one existing file changed
  E. previously approved mapping retained
  F. stronger conflicting SOLD evidence surfaced for review
  G. no duplicate evidence after repeated unchanged runs
"""
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from openpyxl.styles import PatternFill

from app.update_master import run_update_master
from app.config import MatcherConfig
from app import local_index


def _make_quotation(path, db_code, factory_code, price=10.0, sheet_headers=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    if sheet_headers:
        ws.cell(row=1, column=1, value="Item Code")
        ws.cell(row=1, column=3, value="Factory Code")
        ws.cell(row=1, column=3).fill = yellow
        ws.cell(row=1, column=4, value="Unit Price (USD EXW)")
    ws.cell(row=2, column=1, value=db_code)
    ws.cell(row=2, column=2, value="plate")
    ws.cell(row=2, column=3, value=factory_code)
    ws.cell(row=2, column=3).fill = yellow
    ws.cell(row=2, column=4, value=price)
    wb.save(path)


def _make_sales_report(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Product | Material Code", "Product | Material Brief Description",
               "Sales Invoice (Product) Quantity", "Sales Invoice (Product) Unit Price",
               "Sales Invoice Date", "Customer Code"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for i, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=v)
    wb.save(path)


class UpdateMasterTestBase(unittest.TestCase):
    def setUp(self):
        self.root = Path(tempfile.mkdtemp())
        self.factory_dir = self.root / "factory"
        self.customer_dir = self.root / "customer"
        self.factory_dir.mkdir()
        self.customer_dir.mkdir()
        self.index_path = self.root / "master_index.db"
        self.cfg = MatcherConfig()


class TestA_FirstRun(UpdateMasterTestBase):
    def test_first_run_processes_all_files_as_new(self):
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")

        result = run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)
        self.assertEqual(result.stats.files_scanned, 2)
        self.assertEqual(result.stats.new_processed, 2)
        self.assertEqual(result.stats.changed_reprocessed, 0)
        self.assertEqual(result.stats.unchanged_skipped, 0)
        self.assertGreater(result.stats.new_evidence_added, 0)
        self.assertEqual(local_index.processed_file_count(self.index_path), 2)


class TestB_SecondRunNoChanges(UpdateMasterTestBase):
    def test_immediate_second_run_skips_everything(self):
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")

        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)
        result2 = run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        self.assertEqual(result2.stats.new_processed, 0)
        self.assertEqual(result2.stats.changed_reprocessed, 0)
        self.assertEqual(result2.stats.unchanged_skipped, 2)
        self.assertEqual(result2.stats.new_evidence_added, 0)


class TestC_OneNewFileAdded(UpdateMasterTestBase):
    def test_new_file_added_between_runs_is_processed_alone(self):
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")
        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        _make_quotation(self.factory_dir / "f2.xlsx", "DB200002", "H0002")
        result2 = run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        self.assertEqual(result2.stats.new_processed, 1)
        self.assertEqual(result2.stats.changed_reprocessed, 0)
        self.assertEqual(result2.stats.unchanged_skipped, 2)  # the two original files, untouched


class TestD_ExistingFileChanged(UpdateMasterTestBase):
    def test_changed_file_is_reprocessed_and_old_evidence_replaced(self):
        f1 = self.factory_dir / "f1.xlsx"
        _make_quotation(f1, "DB100001", "H0001", price=10.0)
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001", price=20.0)
        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)
        evidence_before = local_index.load_all_evidence(self.index_path)
        prices_before = {r.unit_price for r in evidence_before if r.source == "f1.xlsx"}
        self.assertIn(10.0, prices_before)

        # Revise the factory file: same product, corrected price.
        _make_quotation(f1, "DB100001", "H0001", price=15.0)
        result2 = run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        self.assertEqual(result2.stats.changed_reprocessed, 1)
        self.assertEqual(result2.stats.new_processed, 0)
        evidence_after = local_index.load_all_evidence(self.index_path)
        f1_records = [r for r in evidence_after if r.source == "f1.xlsx"]
        self.assertEqual(len(f1_records), 1)  # old contribution replaced, not duplicated
        self.assertEqual(f1_records[0].unit_price, 15.0)  # the corrected price


class TestE_ApprovedMappingRetained(UpdateMasterTestBase):
    def test_approved_mapping_survives_a_later_run(self):
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")
        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        local_index.mark_mapping_reviewed("H0001", "DB100001", "confirmed correct", self.index_path)

        # Add an unrelated new file -- the approved mapping must survive untouched.
        _make_quotation(self.factory_dir / "f2.xlsx", "DB200002", "H0002")
        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        self.assertEqual(local_index.get_reviewed_mapping("H0001", self.index_path), "DB100001")


class TestF_StrongerSoldEvidenceSurfacedNotOverriding(UpdateMasterTestBase):
    def test_conflicting_sold_evidence_surfaces_for_review_without_overwriting(self):
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")
        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)
        local_index.mark_mapping_reviewed("H0001", "DB100001", "confirmed correct", self.index_path)

        # A real sale under a DIFFERENT DB code for the same factory part.
        sales_path = self.root / "sales.xlsx"
        _make_sales_report(sales_path, [("DB999999", "plate", 100, 5.0, "2026-06-01", "Hotel X")])
        self.cfg.sales_report_path = str(sales_path)
        # Cross-reference needs H0001 already known as DB999999's factory code --
        # simulate this by also quoting DB999999 under H0001 in a factory file.
        _make_quotation(self.factory_dir / "f3.xlsx", "DB999999", "H0001")

        result2 = run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        # The approved mapping must NOT have changed.
        self.assertEqual(local_index.get_reviewed_mapping("H0001", self.index_path), "DB100001")
        # But the conflict must be surfaced.
        self.assertGreater(result2.stats.mappings_requiring_review, 0)
        conflict_items = [it for it in result2.review_items if it.factory_code == "H0001"]
        self.assertTrue(any(it.prior_approved_mapping == "DB100001" for it in conflict_items))
        sold_conflict = [it for it in conflict_items if it.evidence_tier == "SOLD"]
        self.assertTrue(sold_conflict)
        self.assertEqual(sold_conflict[0].confidence, "High")


class TestG_NoDuplicateEvidenceAfterRepeatedUnchangedRuns(UpdateMasterTestBase):
    def test_repeated_unchanged_runs_do_not_grow_evidence_store(self):
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")

        run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)
        count_after_first = local_index.evidence_count(self.index_path)

        for _ in range(3):
            run_update_master([self.factory_dir], [self.customer_dir], self.cfg, self.index_path)

        self.assertEqual(local_index.evidence_count(self.index_path), count_after_first)


class TestAnalysePathUnaffected(UpdateMasterTestBase):
    """Sanity check: run_matching() (the existing Analyse path) must
    remain completely functional and untouched by any of this."""

    def test_run_matching_still_works_independently(self):
        from app.orchestrator import run_matching
        _make_quotation(self.factory_dir / "f1.xlsx", "DB100001", "H0001")
        _make_quotation(self.customer_dir / "c1.xlsx", "DB100001", "H0001")
        result = run_matching(self.factory_dir, self.customer_dir, self.cfg, photo_store_dir=self.root / "photo_store")
        self.assertEqual(result.stats.factory_files_found, 1)
        self.assertEqual(result.stats.customer_files_found, 1)


class TestReviewItemUniqueIdentity(unittest.TestCase):
    """Regression: two field-for-field identical ReviewItems must each
    have a distinct, stable item_id, and removing/acting on one by its
    item_id must never affect the other -- this is the exact edge case
    identified with dataclass-value-based removal (fixed by giving
    each ReviewItem its own uuid-based item_id instead)."""

    def test_identical_review_items_get_distinct_ids(self):
        from app.update_master import ReviewItem
        a = ReviewItem(
            factory_code="H0001", proposed_db_code="DB100001", finish_code="",
            description="plate", source="f.xlsx", evidence_tier="QUOTED",
            confidence="Medium", prior_approved_mapping="", reason="conflict",
        )
        b = ReviewItem(
            factory_code="H0001", proposed_db_code="DB100001", finish_code="",
            description="plate", source="f.xlsx", evidence_tier="QUOTED",
            confidence="Medium", prior_approved_mapping="", reason="conflict",
        )
        # Field-for-field identical in every OTHER way...
        self.assertEqual(a.factory_code, b.factory_code)
        self.assertEqual(a.proposed_db_code, b.proposed_db_code)
        self.assertEqual(a.description, b.description)
        # ...but their identifiers must differ.
        self.assertNotEqual(a.item_id, b.item_id)

    def test_removing_one_of_two_identical_items_by_id_leaves_the_other(self):
        from app.update_master import ReviewItem

        def make_identical():
            return ReviewItem(
                factory_code="H0001", proposed_db_code="DB100001", finish_code="",
                description="plate", source="f.xlsx", evidence_tier="QUOTED",
                confidence="Medium", prior_approved_mapping="", reason="conflict",
            )

        a, b = make_identical(), make_identical()
        all_items = [a, b]

        # Simulate the GUI's fixed removal logic directly (the same
        # id-based filter used in gui.py's _remove_review_rows).
        remove_ids = {a.item_id}
        remaining = [it for it in all_items if it.item_id not in remove_ids]

        self.assertEqual(len(remaining), 1)
        self.assertEqual(remaining[0].item_id, b.item_id)  # only 'b' survives
        self.assertNotIn(a.item_id, {it.item_id for it in remaining})


if __name__ == "__main__":
    unittest.main()
