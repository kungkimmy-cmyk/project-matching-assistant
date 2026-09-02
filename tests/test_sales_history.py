import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from app.config import MatcherConfig
from app.sales_history import read_sales_history


def _make_sales_report(path, rows, title_block_rows=2):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, title_block_rows + 1):
        ws.cell(row=r, column=1, value=f"Mimosa Sales Invoice Report - generated {r}")
    header_row = title_block_rows + 1
    headers = ["Product | Material Code", "Product | Material Brief Description",
               "Sales Invoice (Product) Quantity", "Sales Invoice (Product) Unit Price",
               "Sales Invoice Date", "Customer Code"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    for i, row in enumerate(rows, start=header_row + 1):
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=v)
    wb.save(path)


class TestReadSalesHistory(unittest.TestCase):
    def setUp(self):
        self.cfg = MatcherConfig()

    def test_missing_file_returns_empty_not_crash(self):
        result = read_sales_history("/nonexistent/path/sales.xlsx", self.cfg)
        self.assertEqual(result, [])

    def test_reads_records_past_title_block(self):
        path = Path(tempfile.mktemp(suffix=".xlsx"))
        _make_sales_report(path, [
            ("DB30H0025", "9 inch plate", 100, 3.92, "2026-01-15", "Hotel A"),
            ("DB30H0023", "8 inch plate", 50, 2.23, "2026-02-01", "Hotel B"),
        ])
        records = read_sales_history(path, self.cfg)
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].db_code, "DB30H0025")
        self.assertEqual(records[0].quantity, 100)
        self.assertEqual(records[0].customer, "Hotel A")
        self.assertIsNone(records[0].factory_code)  # real report has no factory-code column

    def test_rows_missing_both_codes_skipped(self):
        path = Path(tempfile.mktemp(suffix=".xlsx"))
        _make_sales_report(path, [
            (None, "some note row", None, None, None, None),
            ("DB30H0025", "9 inch plate", 100, 3.92, "2026-01-15", "Hotel A"),
        ])
        records = read_sales_history(path, self.cfg)
        self.assertEqual(len(records), 1)

    def test_corrupted_file_returns_empty_not_crash(self):
        path = Path(tempfile.mktemp(suffix=".xlsx"))
        path.write_text("not a real xlsx file")
        result = read_sales_history(path, self.cfg)
        self.assertEqual(result, [])


REAL_SALES_REPORT_PATH = Path("/mnt/user-data/uploads/Sales_Invoice_Report_20260902Sales_Invoice_Report.xlsx")


@unittest.skipUnless(REAL_SALES_REPORT_PATH.exists(), "Real sales report not present in this environment")
class TestRealSalesReport(unittest.TestCase):
    """Integration test against the actual uploaded Mimosa Sales
    Invoice Report -- not a synthetic fixture. Confirms the read_only
    dimension bug fix and the real column mapping both work end-to-end."""

    @classmethod
    def setUpClass(cls):
        cls.cfg = MatcherConfig(sales_report_path=str(REAL_SALES_REPORT_PATH))
        cls.records = read_sales_history(REAL_SALES_REPORT_PATH, cls.cfg)

    def test_real_file_produces_a_substantial_number_of_records(self):
        # The real file has 1,629 data rows -- if the read_only
        # dimension bug regresses, this drops to 0.
        self.assertGreater(len(self.records), 1000)

    def test_real_records_have_db_codes(self):
        with_db_code = [r for r in self.records if r.db_code]
        self.assertGreater(len(with_db_code), 1000)

    def test_real_records_have_no_factory_code(self):
        # Confirmed structural fact about this report -- if this ever
        # fails, the report layout changed and needs re-validating.
        self.assertTrue(all(r.factory_code is None for r in self.records))

    def test_real_records_have_dates(self):
        with_dates = [r for r in self.records if r.date]
        self.assertGreater(len(with_dates), 1000)

    def test_real_records_have_quantities_and_prices(self):
        with_qty = [r for r in self.records if r.quantity is not None]
        with_price = [r for r in self.records if r.unit_price is not None]
        self.assertGreater(len(with_qty), 1000)
        self.assertGreater(len(with_price), 1000)


if __name__ == "__main__":
    unittest.main()
